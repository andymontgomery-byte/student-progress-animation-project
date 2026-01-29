#!/usr/bin/env python3
"""
Build student progress table from MAP data and norms tables.
"""

import csv
import os
from collections import defaultdict
from openpyxl import load_workbook

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
FALL_FILE = os.path.join(BASE_DIR, 'data', 'input', 'fall_data.xlsx')
WINTER_FILE = os.path.join(BASE_DIR, 'data', 'input', 'winter_data.xlsx')
NORMS_FILE = os.path.join(BASE_DIR, 'norms_tables', 'csv', 'student_status_percentiles_2025.csv')
OUTPUT_FILE = os.path.join(BASE_DIR, 'data', 'output', 'student_progress.csv')

# Map course names from data to subject names in norms
COURSE_TO_SUBJECT = {
    'Math K-12': 'Mathematics',
    'Reading': 'Reading',
    'Language Usage': 'Language Usage',
    'Science K-12': 'Science'
}

def load_norms():
    """Load norms table into lookup dict: (subject, term, grade, percentile) -> rit_score"""
    norms = {}
    with open(NORMS_FILE) as f:
        reader = csv.DictReader(f)
        for row in reader:
            key = (row['subject'], row['term'], row['grade'], int(row['percentile']))
            norms[key] = int(row['rit_score'])
    return norms

def load_norms_by_rit():
    """Load norms grouped by (subject, term, grade) with rit->percentile mapping"""
    norms_by_group = defaultdict(dict)
    with open(NORMS_FILE) as f:
        reader = csv.DictReader(f)
        for row in reader:
            group_key = (row['subject'], row['term'], row['grade'])
            rit = int(row['rit_score'])
            pct = int(row['percentile'])
            # Store the percentile for this RIT
            # If multiple percentiles have same RIT, keep the HIGHER percentile
            # (student qualifies for the highest percentile at their RIT level)
            if rit not in norms_by_group[group_key] or pct > norms_by_group[group_key][rit]:
                norms_by_group[group_key][rit] = pct
    return norms_by_group

def rit_to_percentile(norms_by_group, subject, term, grade, rit_score):
    """Look up percentile for a RIT score, interpolating as needed"""
    if rit_score is None:
        return None

    group_key = (subject, term, grade)
    if group_key not in norms_by_group:
        return None

    rit_to_pct = norms_by_group[group_key]

    # Direct lookup
    if rit_score in rit_to_pct:
        return rit_to_pct[rit_score]

    # Find surrounding RIT values
    rit_values = sorted(rit_to_pct.keys())

    # Below minimum -> percentile 1
    if rit_score < rit_values[0]:
        return 1

    # Above maximum -> percentile 99
    if rit_score > rit_values[-1]:
        return 99

    # Find the closest RIT that's <= our score
    for i, rit in enumerate(rit_values):
        if rit > rit_score:
            # Use the lower RIT's percentile (conservative)
            return rit_to_pct[rit_values[i-1]]

    return rit_to_pct[rit_values[-1]]

def calculate_99_levels(norms, subject, term, grade, rit_score):
    """Calculate how many grade levels beyond current grade the student is 99th percentile"""
    if rit_score is None or grade is None:
        return 0

    # First check if student is at 99th percentile at their own grade
    key = (subject, term, grade, 99)
    if key not in norms:
        return 0

    if rit_score < norms[key]:
        return 0  # Not 99th at own grade

    # Count how many grades above they're still 99th percentile
    levels = 0

    # Get ordered list of grades
    grade_order = ['K'] + [str(i) for i in range(1, 13)]

    try:
        current_idx = grade_order.index(grade)
    except ValueError:
        return 0

    for higher_grade in grade_order[current_idx + 1:]:
        key = (subject, term, higher_grade, 99)
        if key in norms and rit_score >= norms[key]:
            levels += 1
        else:
            break  # Stop at first grade where they're not 99th

    return levels

def load_fall_data():
    """Load fall data from Excel"""
    wb = load_workbook(FALL_FILE)
    ws = wb.active

    # Find header row (row 3 in this file)
    headers = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if row[0] == 'email':
            headers = row
            header_row = i
            break

    if not headers:
        raise ValueError("Could not find headers in fall data")

    data = {}
    for row in ws.iter_rows(min_row=header_row + 2, values_only=True):
        if not row[0]:  # Skip empty rows
            continue

        record = dict(zip(headers, row))
        email = record.get('email')
        course = record.get('course')

        if not email or not course:
            continue

        key = (email, course)
        data[key] = {
            'email': email,
            'course': course,
            'grade': record.get('grade'),
            'school': record.get('schoolname'),
            'fall_rit': record.get('testritscore'),
            'fall_pct': record.get('testpercentile')
        }

    return data

def load_winter_data():
    """Load winter data from Excel"""
    wb = load_workbook(WINTER_FILE)
    ws = wb.active

    # Find header row
    headers = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if row[0] == 'email':
            headers = row
            header_row = i
            break

    if not headers:
        raise ValueError("Could not find headers in winter data")

    data = {}
    for row in ws.iter_rows(min_row=header_row + 2, values_only=True):
        if not row[0]:
            continue

        record = dict(zip(headers, row))
        email = record.get('email')
        course = record.get('course')

        if not email or not course:
            continue

        key = (email, course)
        data[key] = {
            'email': email,
            'course': course,
            'grade': record.get('grade'),
            'school': record.get('schoolname'),
            'winter_rit': record.get('testritscore'),
            'winter_pct': record.get('testpercentile'),
            'cgi': record.get('falltowinterconditionalgrowthindex'),
            'observed_growth': record.get('falltowinterobservedgrowth')
        }

    return data

def build_table():
    """Build the student progress table"""
    print("Loading norms...")
    norms = load_norms()
    norms_by_rit = load_norms_by_rit()

    print("Loading fall data...")
    fall_data = load_fall_data()
    print(f"  Loaded {len(fall_data)} fall records")

    print("Loading winter data...")
    winter_data = load_winter_data()
    print(f"  Loaded {len(winter_data)} winter records")

    print("Building table...")
    results = []

    # Only include students who have winter data (they're the ones with CGI)
    for key, winter in winter_data.items():
        email, course = key

        # Get fall data for this student/course
        fall = fall_data.get(key, {})

        # Skip courses that don't map to norms
        subject = COURSE_TO_SUBJECT.get(course)
        if not subject:
            print(f"  Warning: Skipping unknown course '{course}' for {email}")
            continue

        # Get grade (from winter data, fall back to fall data)
        grade = winter.get('grade') or fall.get('grade')
        if not grade:
            print(f"  Warning: No grade for {email} {course}")
            continue

        # Convert grade to string for lookup
        grade = str(grade)
        if grade == '0':
            grade = 'K'

        # Get RIT scores
        fall_rit = fall.get('fall_rit')
        winter_rit = winter.get('winter_rit')

        # Get percentiles from source data
        fall_pct = fall.get('fall_pct')
        winter_pct = winter.get('winter_pct')

        # Convert to int if not None
        if fall_pct is not None:
            fall_pct = int(fall_pct)
        if winter_pct is not None:
            winter_pct = int(winter_pct)

        # Calculate projected RIT
        observed_growth = winter.get('observed_growth')
        projected_rit = None
        if winter_rit is not None and observed_growth is not None:
            projected_rit = int(winter_rit) + int(observed_growth)

        # Look up projected percentile from Spring norms
        projected_pct = rit_to_percentile(norms_by_rit, subject, 'Spring', grade, projected_rit)

        # Calculate 99th grade levels using 2025 norms
        fall_99_levels = calculate_99_levels(norms, subject, 'Fall', grade,
                                              int(fall_rit) if fall_rit else None)
        winter_99_levels = calculate_99_levels(norms, subject, 'Winter', grade,
                                                int(winter_rit) if winter_rit else None)
        projected_99_levels = calculate_99_levels(norms, subject, 'Spring', grade, projected_rit)

        # Get CGI
        cgi = winter.get('cgi')
        if cgi is not None:
            cgi = round(float(cgi), 2)

        results.append({
            'email': email,
            'grade': grade,
            'school': winter.get('school') or fall.get('school'),
            'course': course,
            'fall_pct': fall_pct,
            'fall_99_levels': fall_99_levels,
            'winter_pct': winter_pct,
            'winter_99_levels': winter_99_levels,
            'projected_pct': projected_pct,
            'projected_99_levels': projected_99_levels,
            'cgi': cgi
        })

    # Sort by email, then course
    results.sort(key=lambda r: (r['email'], r['course']))

    print(f"  Built {len(results)} records")

    # Write output
    print(f"Writing to {OUTPUT_FILE}...")
    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)

    fieldnames = ['email', 'grade', 'school', 'course',
                  'fall_pct', 'fall_99_levels',
                  'winter_pct', 'winter_99_levels',
                  'projected_pct', 'projected_99_levels', 'cgi']

    with open(OUTPUT_FILE, 'w', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(results)

    print("Done!")

    # Summary
    high_cgi = len([r for r in results if r['cgi'] and r['cgi'] > 0.8])
    print(f"\nSummary:")
    print(f"  Total records: {len(results)}")
    print(f"  High CGI (>0.8): {high_cgi}")

    return results

if __name__ == '__main__':
    build_table()
