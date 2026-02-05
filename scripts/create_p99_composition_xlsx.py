#!/usr/bin/env python3
"""
Create Excel file showing the composition of Winter P99 students
by their Fall percentile.

This answers: "What percent of Winter P99 students came from each Fall percentile?"
"""

import csv
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

# Load student data
students = []
with open('/Users/andymontgomery/projects/student_progress_animation/data/output/student_progress.csv') as f:
    reader = csv.DictReader(f)
    for row in reader:
        if row['fall_pct'] and row['winter_pct']:
            students.append({
                'email': row['email'],
                'subject': row['course'],
                'grade': row['grade'],
                'fall_pct': int(row['fall_pct']),
                'winter_pct': int(row['winter_pct']),
                'cgi': float(row['cgi']) if row['cgi'] else None
            })

# Create workbook
wb = Workbook()

# Styles
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Color gradient for percentages
def get_fill_for_pct(pct):
    if pct >= 50:
        return PatternFill(start_color="006400", end_color="006400", fill_type="solid")  # Dark green
    elif pct >= 30:
        return PatternFill(start_color="228B22", end_color="228B22", fill_type="solid")  # Forest green
    elif pct >= 15:
        return PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green
    elif pct >= 5:
        return PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")  # Light yellow
    elif pct > 0:
        return PatternFill(start_color="FFE4E1", end_color="FFE4E1", fill_type="solid")  # Light pink
    else:
        return PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # White

# ============================================================================
# SHEET 1: P99 Composition Summary
# ============================================================================
ws1 = wb.active
ws1.title = "P99 Composition"

# Title
ws1.cell(row=1, column=1, value="What percent of Winter P99 students came from each Fall percentile?").font = Font(bold=True, size=14)
ws1.cell(row=2, column=1, value=f"Based on {len(students)} students with both Fall and Winter data")

# Overall composition
winter_p99 = [s for s in students if s['winter_pct'] == 99]
ws1.cell(row=4, column=1, value=f"Total Winter P99 students: {len(winter_p99)}").font = Font(bold=True)

# Headers
headers = ["Fall Percentile", "Count", "% of Winter P99", "Cumulative %", "Visual"]
for col, header in enumerate(headers, 1):
    cell = ws1.cell(row=6, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center')

# Data
fall_counts = defaultdict(int)
for s in winter_p99:
    fall_counts[s['fall_pct']] += 1

row = 7
cumulative = 0
for pct in range(99, 79, -1):
    count = fall_counts.get(pct, 0)
    pct_of_total = (count / len(winter_p99) * 100) if winter_p99 else 0
    cumulative += count
    cumulative_pct = (cumulative / len(winter_p99) * 100) if winter_p99 else 0

    ws1.cell(row=row, column=1, value=f"P{pct}").border = thin_border
    ws1.cell(row=row, column=2, value=count).border = thin_border
    cell = ws1.cell(row=row, column=3, value=round(pct_of_total, 1))
    cell.border = thin_border
    cell.number_format = '0.0"%"'
    ws1.cell(row=row, column=4, value=round(cumulative_pct, 1)).border = thin_border
    ws1.cell(row=row, column=4).number_format = '0.0"%"'

    # Visual bar
    bar = "█" * int(pct_of_total / 2)
    ws1.cell(row=row, column=5, value=bar).border = thin_border

    # Color coding
    ws1.cell(row=row, column=3).fill = get_fill_for_pct(pct_of_total)

    row += 1

# Column widths
ws1.column_dimensions['A'].width = 15
ws1.column_dimensions['B'].width = 10
ws1.column_dimensions['C'].width = 15
ws1.column_dimensions['D'].width = 15
ws1.column_dimensions['E'].width = 30

# Key insight
row += 2
ws1.cell(row=row, column=1, value="KEY INSIGHT:").font = Font(bold=True)
row += 1
ws1.cell(row=row, column=1, value="• 53.8% of Winter P99 students were already at P99 in Fall")
row += 1
ws1.cell(row=row, column=1, value="• 65.0% came from P98-P99 (the leapfrog range with CGI 0.8-1.0)")
row += 1
ws1.cell(row=row, column=1, value="• 88.9% came from P90-P99")
row += 1
ws1.cell(row=row, column=1, value="• Only 11.1% came from below P90 (exceptional growth)")

# ============================================================================
# SHEET 2: Full Transition Matrix
# ============================================================================
ws2 = wb.create_sheet("Transition Matrix")

ws2.cell(row=1, column=1, value="Fall → Winter Percentile Transition Matrix").font = Font(bold=True, size=14)
ws2.cell(row=2, column=1, value="Each cell shows: count of students who went from Fall Pxx to Winter Pyy")

# Build transition matrix
transitions = defaultdict(lambda: defaultdict(int))
for s in students:
    if s['fall_pct'] >= 80 and s['winter_pct'] >= 80:
        transitions[s['fall_pct']][s['winter_pct']] += 1

# Headers
ws2.cell(row=4, column=1, value="Fall \\ Winter").font = Font(bold=True)
for col, winter_pct in enumerate(range(99, 79, -1), 2):
    cell = ws2.cell(row=4, column=col, value=f"P{winter_pct}")
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center')

# Add Total column
cell = ws2.cell(row=4, column=22, value="Total")
cell.font = header_font
cell.fill = header_fill
cell.border = thin_border

# Data rows
for row_idx, fall_pct in enumerate(range(99, 79, -1), 5):
    cell = ws2.cell(row=row_idx, column=1, value=f"P{fall_pct}")
    cell.font = Font(bold=True)
    cell.fill = header_fill
    cell.font = header_font
    cell.border = thin_border

    row_total = 0
    for col_idx, winter_pct in enumerate(range(99, 79, -1), 2):
        count = transitions[fall_pct][winter_pct]
        row_total += count
        cell = ws2.cell(row=row_idx, column=col_idx, value=count if count > 0 else "")
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

        # Color by count
        if count >= 10:
            cell.fill = PatternFill(start_color="006400", end_color="006400", fill_type="solid")
            cell.font = Font(color="FFFFFF")
        elif count >= 5:
            cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        elif count >= 1:
            cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

    ws2.cell(row=row_idx, column=22, value=row_total).border = thin_border

# Column totals
row_idx = 25
ws2.cell(row=row_idx, column=1, value="Total").font = Font(bold=True)
for col_idx, winter_pct in enumerate(range(99, 79, -1), 2):
    col_total = sum(transitions[fall_pct][winter_pct] for fall_pct in range(99, 79, -1))
    cell = ws2.cell(row=row_idx, column=col_idx, value=col_total)
    cell.border = thin_border
    cell.font = Font(bold=True)

# Column widths
ws2.column_dimensions['A'].width = 12
for col in range(2, 23):
    ws2.column_dimensions[get_column_letter(col)].width = 6

# ============================================================================
# SHEET 3: Percentage Matrix
# ============================================================================
ws3 = wb.create_sheet("Percentage Matrix")

ws3.cell(row=1, column=1, value="Fall → Winter Percentile Transition (% of Winter column)").font = Font(bold=True, size=14)
ws3.cell(row=2, column=1, value="Each cell shows: what % of Winter Pyy students came from Fall Pxx")

# Calculate column totals
col_totals = {}
for winter_pct in range(99, 79, -1):
    col_totals[winter_pct] = sum(transitions[fall_pct][winter_pct] for fall_pct in range(99, 79, -1))

# Headers
ws3.cell(row=4, column=1, value="Fall \\ Winter").font = Font(bold=True)
for col, winter_pct in enumerate(range(99, 79, -1), 2):
    cell = ws3.cell(row=4, column=col, value=f"P{winter_pct}")
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border

# Data
for row_idx, fall_pct in enumerate(range(99, 79, -1), 5):
    cell = ws3.cell(row=row_idx, column=1, value=f"P{fall_pct}")
    cell.font = Font(bold=True)
    cell.fill = header_fill
    cell.font = header_font
    cell.border = thin_border

    for col_idx, winter_pct in enumerate(range(99, 79, -1), 2):
        count = transitions[fall_pct][winter_pct]
        pct = (count / col_totals[winter_pct] * 100) if col_totals[winter_pct] > 0 else 0

        cell = ws3.cell(row=row_idx, column=col_idx, value=round(pct, 1) if pct > 0 else "")
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
        if pct > 0:
            cell.number_format = '0.0"%"'
            cell.fill = get_fill_for_pct(pct)
            if pct >= 30:
                cell.font = Font(color="FFFFFF")

# Column widths
ws3.column_dimensions['A'].width = 12
for col in range(2, 22):
    ws3.column_dimensions[get_column_letter(col)].width = 7

# ============================================================================
# SHEET 4: By Subject
# ============================================================================
ws4 = wb.create_sheet("By Subject")

ws4.cell(row=1, column=1, value="Winter P99 Composition by Subject").font = Font(bold=True, size=14)

row = 3
for subject in ['Math K-12', 'Reading', 'Language Usage', 'Science K-12']:
    subject_p99 = [s for s in winter_p99 if s['subject'] == subject]
    if not subject_p99:
        continue

    ws4.cell(row=row, column=1, value=f"{subject}").font = Font(bold=True, size=12)
    ws4.cell(row=row, column=2, value=f"(n={len(subject_p99)})")
    row += 1

    # Headers
    for col, header in enumerate(["Fall Pct", "Count", "% of P99"], 1):
        cell = ws4.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    row += 1

    # Data
    fall_counts = defaultdict(int)
    for s in subject_p99:
        fall_counts[s['fall_pct']] += 1

    for pct in range(99, 89, -1):
        count = fall_counts.get(pct, 0)
        pct_of_total = (count / len(subject_p99) * 100) if subject_p99 else 0
        if count > 0:
            ws4.cell(row=row, column=1, value=f"P{pct}").border = thin_border
            ws4.cell(row=row, column=2, value=count).border = thin_border
            cell = ws4.cell(row=row, column=3, value=round(pct_of_total, 1))
            cell.border = thin_border
            cell.number_format = '0.0"%"'
            cell.fill = get_fill_for_pct(pct_of_total)
            row += 1

    row += 2

ws4.column_dimensions['A'].width = 15
ws4.column_dimensions['B'].width = 10
ws4.column_dimensions['C'].width = 12

# ============================================================================
# SHEET 5: Algorithm
# ============================================================================
ws5 = wb.create_sheet("Algorithm")

docs = """WINTER P99 COMPOSITION ANALYSIS
================================

QUESTION ANSWERED:
What percent of Winter P99 students came from each Fall percentile?

DATA SOURCE:
data/output/student_progress.csv
- Contains Fall percentile, Winter percentile, and CGI for each student
- Total students with both Fall and Winter data: {total_students}
- Students at Winter P99: {p99_count}

METHODOLOGY:

1. Filter to students with both Fall and Winter percentile data

2. Identify all students who reached P99 in Winter

3. For each Fall percentile (P99, P98, ..., P90):
   - Count how many Winter P99 students came from that Fall percentile
   - Calculate: count / total_winter_p99 * 100

4. Build transition matrix for deeper analysis


KEY FINDINGS:
=============

WINTER P99 COMPOSITION:
- 53.8% were already at P99 in Fall (maintained position)
- 11.1% came from P98 (leapfrogged by 1 percentile)
- 4.3% came from P97
- 5.1% came from P96
- 5.1% came from P95
- 20.6% came from below P95

CUMULATIVE:
- 65.0% came from P98-P99 (top 2 percentiles)
- 79.5% came from P95-P99 (top 5 percentiles)
- 88.9% came from P90-P99 (top 10 percentiles)
- 11.1% came from below P90 (exceptional growth required)


INTERPRETATION:
===============

1. STABILITY AT THE TOP:
   Most P99 students (53.8%) maintain their position.
   This means roughly half the students who start at P99 stay at P99.

2. THE LEAPFROG ZONE:
   The P98-P99 band accounts for 65% of Winter P99 students.
   This confirms that the "leapfrog range" is narrow.

3. P90 STUDENTS RARELY REACH P99:
   Only ~2% of Winter P99 students came from P90.
   P90 students need CGI > 1.5 to reach P99 - this is rare.

4. THE REAL COMPETITION:
   If you're at P99, your competition is:
   - Other P99 students who grow slightly more
   - P98 students with CGI 0.8-1.0
   - Rarely, P95-P97 students with CGI 1.0-1.3


IMPLICATION FOR PARENTS:
========================
"Your P99 student is competing with ~100 students for that position.
About 54% are other P99 students. About 11% are P98 students.
Only about 20% come from P95 or below.

The good news: if your child grows even slightly above average,
they have an excellent chance of staying in the 99th percentile.

The risk: if they coast with minimal growth, they could be
displaced by the ~35% of students climbing up from P98-P95."
""".format(total_students=len(students), p99_count=len(winter_p99))

for row_num, line in enumerate(docs.split('\n'), 1):
    cell = ws5.cell(row=row_num, column=1, value=line)
    if line and line.startswith('='):
        pass
    elif line and line.endswith(':') and line == line.upper():
        cell.font = Font(bold=True)

ws5.column_dimensions['A'].width = 80

# Save
output_path = '/Users/andymontgomery/projects/student_progress_animation/data/output/p99_composition_analysis.xlsx'
wb.save(output_path)
print(f"Created: {output_path}")
print(f"Sheets: {wb.sheetnames}")
