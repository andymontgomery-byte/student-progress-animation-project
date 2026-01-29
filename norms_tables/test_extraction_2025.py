#!/usr/bin/env python3
"""
Test suite for 2025 norms table extraction from Excel.
Run: python3 norms_tables/test_extraction_2025.py
"""

from openpyxl import load_workbook
import csv
import os
import random
import sys

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = os.path.join(BASE_DIR, 'norms_2025.xlsx')
CSV_FILE = os.path.join(BASE_DIR, 'csv', 'student_status_percentiles_2025.csv')

SHEET_TO_SUBJECT = {
    'Math': 'Mathematics',
    'Reading': 'Reading',
    'Language Usage': 'Language Usage',
    'Science': 'Science'
}

def load_csv():
    """Load extracted CSV data"""
    data = {}
    with open(CSV_FILE) as f:
        reader = csv.DictReader(f)
        for row in reader:
            key = (row['subject'], row['term'], row['grade'], int(row['percentile']))
            data[key] = int(row['rit_score'])
    return data

def load_excel_values():
    """Load values directly from Excel"""
    wb = load_workbook(EXCEL_FILE)
    values = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        subject = SHEET_TO_SUBJECT[sheet_name]

        terms = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
        grades = list(ws.iter_rows(min_row=3, max_row=3, values_only=True))[0]

        for row in ws.iter_rows(min_row=8, values_only=True):
            if row[0] and str(row[0]).startswith('P'):
                pct = int(row[0][1:])
                for col_idx in range(1, len(row)):
                    if terms[col_idx] and grades[col_idx] and row[col_idx]:
                        term = terms[col_idx].capitalize()
                        grade = str(grades[col_idx])
                        if grade == '0':
                            grade = 'K'
                        rit = int(row[col_idx])
                        values.append((subject, term, grade, pct, rit))

    return values

def test_csv_exists():
    """Test that CSV file exists"""
    assert os.path.exists(CSV_FILE), f"CSV file not found: {CSV_FILE}"
    print("✓ CSV file exists")

def test_all_subjects_present():
    """Test that all 4 subjects are present"""
    csv_data = load_csv()
    subjects = set(k[0] for k in csv_data.keys())
    expected = {'Mathematics', 'Reading', 'Language Usage', 'Science'}

    assert subjects == expected, f"Expected {expected}, got {subjects}"
    print("✓ All 4 subjects present")

def test_all_terms_present():
    """Test that all 3 terms are present"""
    csv_data = load_csv()
    terms = set(k[1] for k in csv_data.keys())
    expected = {'Fall', 'Winter', 'Spring'}

    assert terms == expected, f"Expected {expected}, got {terms}"
    print("✓ All 3 terms present")

def test_percentiles_1_to_99():
    """Test that percentiles 1-99 are present"""
    csv_data = load_csv()
    percentiles = set(k[3] for k in csv_data.keys())
    expected = set(range(1, 100))

    assert percentiles == expected, f"Missing percentiles: {expected - percentiles}"
    print("✓ All percentiles 1-99 present")

def test_grade_coverage():
    """Test correct grade coverage for each subject"""
    csv_data = load_csv()

    subject_grades = {}
    for key in csv_data.keys():
        subject = key[0]
        grade = key[2]
        if subject not in subject_grades:
            subject_grades[subject] = set()
        subject_grades[subject].add(grade)

    expected = {
        'Mathematics': {str(i) for i in range(1, 13)},      # 1-12
        'Reading': {str(i) for i in range(1, 13)},          # 1-12
        'Language Usage': {str(i) for i in range(2, 12)},   # 2-11
        'Science': {str(i) for i in range(2, 11)},          # 2-10
    }

    for subject, exp_grades in expected.items():
        actual = subject_grades.get(subject, set())
        assert actual == exp_grades, f"{subject}: expected {sorted(exp_grades)}, got {sorted(actual)}"

    print("✓ Grade coverage correct for all subjects")

def test_random_samples_match_excel():
    """Test 100 random samples match Excel source"""
    csv_data = load_csv()
    excel_values = load_excel_values()

    random.seed(42)
    samples = random.sample(excel_values, min(100, len(excel_values)))

    errors = []
    for subject, term, grade, pct, excel_rit in samples:
        csv_key = (subject, term, grade, pct)
        csv_rit = csv_data.get(csv_key)

        if csv_rit != excel_rit:
            errors.append(f"{subject} {term} Grade {grade} P{pct}: Excel={excel_rit}, CSV={csv_rit}")

    assert not errors, f"Mismatches:\n" + "\n".join(errors[:10])
    print(f"✓ 100 random samples match Excel source")

def test_rit_scores_valid_range():
    """Test all RIT scores are in valid range"""
    csv_data = load_csv()

    invalid = [(k, v) for k, v in csv_data.items() if not (100 <= v <= 350)]
    assert not invalid, f"Invalid RIT scores: {invalid[:5]}"
    print("✓ All RIT scores in valid range (100-350)")

def test_rit_scores_monotonic():
    """Test RIT scores increase with percentile"""
    csv_data = load_csv()

    # Group by subject/term/grade
    groups = {}
    for key, rit in csv_data.items():
        group_key = (key[0], key[1], key[2])
        if group_key not in groups:
            groups[group_key] = []
        groups[group_key].append((key[3], rit))  # (percentile, rit)

    errors = []
    for group_key, values in groups.items():
        sorted_vals = sorted(values, key=lambda x: x[0])
        for i in range(1, len(sorted_vals)):
            if sorted_vals[i][1] < sorted_vals[i-1][1]:
                errors.append(f"{group_key}: P{sorted_vals[i-1][0]}={sorted_vals[i-1][1]} > P{sorted_vals[i][0]}={sorted_vals[i][1]}")

    assert not errors, f"Non-monotonic:\n" + "\n".join(errors[:5])
    print("✓ RIT scores monotonically increase with percentile")

def test_row_count():
    """Test expected row count"""
    csv_data = load_csv()

    # Math: 12 grades × 99 percentiles × 3 terms = 3564
    # Reading: 12 grades × 99 percentiles × 3 terms = 3564
    # Language Usage: 10 grades × 99 percentiles × 3 terms = 2970
    # Science: 9 grades × 99 percentiles × 3 terms = 2673
    # Total: 12771

    expected = 12771
    actual = len(csv_data)

    assert actual == expected, f"Expected {expected} rows, got {actual}"
    print(f"✓ Row count correct ({actual})")

def run_all_tests():
    """Run all tests"""
    print("=" * 60)
    print("2025 NORMS EXTRACTION TESTS")
    print("=" * 60)
    print()

    tests = [
        test_csv_exists,
        test_all_subjects_present,
        test_all_terms_present,
        test_percentiles_1_to_99,
        test_grade_coverage,
        test_random_samples_match_excel,
        test_rit_scores_valid_range,
        test_rit_scores_monotonic,
        test_row_count,
    ]

    passed = 0
    failed = 0

    for test in tests:
        try:
            test()
            passed += 1
        except AssertionError as e:
            print(f"✗ {test.__name__}: {e}")
            failed += 1
        except Exception as e:
            print(f"✗ {test.__name__}: ERROR - {e}")
            failed += 1

    print()
    print("=" * 60)
    print(f"RESULTS: {passed} passed, {failed} failed")
    print("=" * 60)

    return failed == 0

if __name__ == '__main__':
    success = run_all_tests()
    sys.exit(0 if success else 1)
